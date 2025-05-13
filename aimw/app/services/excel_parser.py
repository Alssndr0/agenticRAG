import os
from datetime import datetime

import openpyxl
from app.configs.excel_config import get_excel_settings
from app.configs.path_config import get_path_settings
from loguru import logger
from openpyxl.utils import get_column_letter


# Helper function to calculate how many lines a text will occupy
def calculate_line_count(text, chars_per_line):
    if not text:
        return 1

    # Count newlines in the text
    newline_count = text.count("\n")

    # Calculate how many lines the text would wrap to
    remaining_text = text.replace("\n", "")
    remaining_length = len(remaining_text)
    wrap_lines = max(
        1,
        (remaining_length // chars_per_line)
        + (1 if remaining_length % chars_per_line > 0 else 0),
    )

    # Total lines is the sum of explicit newlines and wrapped lines
    return newline_count + wrap_lines


# Helper function to safely get cell color as string
def get_cell_color(cell):
    """Extract cell color as a string, handling different openpyxl versions and formats"""
    if cell.fill.fill_type == "none":
        return None

    if not hasattr(cell.fill.start_color, "rgb") or not cell.fill.start_color.rgb:
        return None

    rgb_value = cell.fill.start_color.rgb

    if isinstance(rgb_value, str):
        # Sometimes the color has 'FF' prefix for alpha channel
        return (
            rgb_value[2:]
            if rgb_value.startswith("FF") and len(rgb_value) == 8
            else rgb_value
        )
    # For RGB objects or other types, try to convert to string
    try:
        # Try getting the hex representation
        rgb_str = str(rgb_value)
        # Remove any non-hex characters and take the last 6 characters
        rgb_str = "".join(c for c in rgb_str if c.upper() in "0123456789ABCDEF")
        return rgb_str[-6:] if len(rgb_str) >= 6 else None
    except Exception:
        return None


def load_workbook_and_get_sheets(excel_file_path, sheet_name=None):
    """
    Loads an Excel workbook and determines which sheets to process

    Parameters:
    excel_file_path (str): Path to the Excel file
    sheet_name (str, optional): Name of specific sheet to process. If None, processes all sheets

    Returns:
    tuple: (workbook, list_of_sheets_to_process)
    """
    workbook = openpyxl.load_workbook(excel_file_path)

    if sheet_name and sheet_name in workbook.sheetnames:
        sheets_to_process = [sheet_name]
    else:
        sheets_to_process = workbook.sheetnames

    return workbook, sheets_to_process


def find_colored_cells_in_row(row):
    """
    Find question and answer cells in a row based on their colors

    Parameters:
    row: Row of cells to process

    Returns:
    tuple: (question_text, question_cell, answer_cell)
    """
    question_text = None
    question_cell = None
    answer_cell = None

    for cell in row:
        # Skip cells with no fill
        if cell.fill.fill_type == "none":
            continue

        if cell_color_str := get_cell_color(cell):
            # Compare with our target colors
            if cell_color_str.upper() == get_excel_settings().QUESTION_COLOR:
                question_text = cell.value
                question_cell = cell
            elif cell_color_str.upper() == get_excel_settings().ANSWER_COLOR:
                answer_cell = cell

    return question_text, question_cell, answer_cell


def process_sheet_questions(worksheet, question_processor_function):
    """
    Processes cells with specific colors, writes answers, and tracks cells for adjustment

    Parameters:
    worksheet (Worksheet): Excel worksheet to process
    question_processor_function (function): Function that takes a question string and returns an answer string

    Returns:
    tuple: (answer_columns, question_columns, max_text_lengths, rows_to_adjust)
    """
    # Track columns containing answer cells for later width adjustment
    answer_columns = set()
    question_columns = set()
    max_text_lengths = {}  # Track max text length per column

    # Track rows that need height adjustment
    rows_to_adjust = {}

    # Process each row in the worksheet
    for row in worksheet.iter_rows():
        # Find question and answer cells
        question_text, question_cell, answer_cell = find_colored_cells_in_row(row)

        # If we found a question but no answer in this row, continue to next row
        if question_text is None or answer_cell is None or question_cell is None:
            continue

        # Track question column
        question_columns.add(question_cell.column)

        # Store row information for height adjustment
        row_idx = answer_cell.row
        if row_idx not in rows_to_adjust:
            rows_to_adjust[row_idx] = {
                "question_cell": question_cell,
                "answer_cell": answer_cell,
            }

        # Process the question and get the answer
        answer_text = question_processor_function(question_text)

        # Write the answer to the answer cell
        answer_cell.value = answer_text

        # Track this column for width adjustment
        col_idx = answer_cell.column
        answer_columns.add(col_idx)

        # Track the maximum text length for this column
        if col_idx not in max_text_lengths:
            max_text_lengths[col_idx] = 0

        # Update max length if this answer is longer
        if answer_text:
            text_length = len(str(answer_text))
            max_text_lengths[col_idx] = max(max_text_lengths[col_idx], text_length)

    return answer_columns, question_columns, max_text_lengths, rows_to_adjust


def apply_text_wrapping_to_column(worksheet, col_idx):
    """
    Applies text wrapping to cells in a column that match the answer color

    Parameters:
    worksheet (Worksheet): Excel worksheet to adjust
    col_idx (int): Column index to apply wrapping
    """
    for row in worksheet.iter_rows():
        if len(row) >= col_idx:
            cell = row[col_idx - 1]  # Adjust for 0-based indexing
            # Get cell color safely
            cell_color = get_cell_color(cell)
            if cell_color and cell_color.upper() == get_excel_settings().ANSWER_COLOR:
                alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
                cell.alignment = alignment


def adjust_column_widths(worksheet, answer_columns, max_text_lengths):
    """
    Adjusts column widths and text wrapping based on content

    Parameters:
    worksheet (Worksheet): Excel worksheet to adjust
    answer_columns (set): Set of column indices that contain answers
    max_text_lengths (dict): Maximum text length for each column
    """
    for col_idx in answer_columns:
        col_letter = get_column_letter(col_idx)

        # Get the maximum text length for this column
        max_length = max_text_lengths.get(col_idx, 0)

        # Calculate a reasonable width based on content length
        # The formula is an approximation: characters × 1.2 for proportional fonts
        # with a minimum width of 15 and maximum of 100
        estimated_width = min(max(15, max_length * 1.2), 100)

        # Apply the width to the column
        worksheet.column_dimensions[col_letter].width = estimated_width

        # Enable text wrapping for very long answers
        if max_length > 80:
            apply_text_wrapping_to_column(worksheet, col_idx)


def get_cell_column_widths(worksheet, question_cell, answer_cell):
    """
    Gets the column widths for question and answer cells

    Parameters:
    worksheet (Worksheet): Excel worksheet
    question_cell: Question cell
    answer_cell: Answer cell

    Returns:
    tuple: (question_column_width, answer_column_width)
    """
    # Get column widths for calculating appropriate height
    q_col_letter = get_column_letter(question_cell.column)
    a_col_letter = get_column_letter(answer_cell.column)

    q_col_width = worksheet.column_dimensions[q_col_letter].width
    a_col_width = worksheet.column_dimensions[a_col_letter].width

    # Default to standard width if not explicitly set
    if not q_col_width:
        q_col_width = 8.43  # Excel default column width
    if not a_col_width:
        a_col_width = 8.43

    return q_col_width, a_col_width


def calculate_cell_line_counts(question_cell, answer_cell, q_col_width, a_col_width):
    """
    Calculates line counts for question and answer cells

    Parameters:
    question_cell: Question cell
    answer_cell: Answer cell
    q_col_width (float): Question column width
    a_col_width (float): Answer column width

    Returns:
    tuple: (question_lines, answer_lines)
    """
    # Calculate approximate character count per line
    q_chars_per_line = int(q_col_width / 1.2)  # ~1.2 chars per unit width
    a_chars_per_line = int(a_col_width / 1.2)

    # Calculate number of lines needed for each cell
    q_text = str(question_cell.value) if question_cell.value else ""
    a_text = str(answer_cell.value) if answer_cell.value else ""

    # Calculate line count for each cell based on word wrapping
    q_lines = calculate_line_count(q_text, q_chars_per_line)
    a_lines = calculate_line_count(a_text, a_chars_per_line)

    return q_lines, a_lines


def adjust_row_heights(worksheet, rows_to_adjust):
    """
    Adjusts row heights based on content in question and answer cells

    Parameters:
    worksheet (Worksheet): Excel worksheet to adjust
    rows_to_adjust (dict): Dictionary mapping row indices to question and answer cells
    """
    for row_idx, cells in rows_to_adjust.items():
        question_cell = cells["question_cell"]
        answer_cell = cells["answer_cell"]

        # Calculate cell dimensions and line counts
        q_col_width, a_col_width = get_cell_column_widths(
            worksheet, question_cell, answer_cell
        )
        q_lines, a_lines = calculate_cell_line_counts(
            question_cell, answer_cell, q_col_width, a_col_width
        )

        # Set the row height based on the maximum number of lines
        max_lines = max(q_lines, a_lines)
        row_height = max(24, max_lines * 14.5)  # Minimum height of 24 points

        # Apply the height to the row
        worksheet.row_dimensions[row_idx].height = row_height

        # Enable text wrapping for both question and answer cells
        alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
        question_cell.alignment = alignment
        answer_cell.alignment = alignment


def ensure_output_directory(output_path):
    """
    Ensures the output directory exists

    Parameters:
    output_path (str): Path to the output directory
    """
    if not os.path.exists(output_path):
        os.makedirs(output_path)


def process_excel_questions(
    excel_file_path,
    question_processor_function,
    sheet_name=None,
    output_dir=get_path_settings().OUTPUT_DIR,
):
    """
    Opens an Excel file, processes cells with specific colors, writes answers,
    and adjusts column widths and row heights for better readability

    Parameters:
    excel_file_path (str): Path to the Excel file
    question_processor_function (function): Function that takes a question string and returns an answer string
    sheet_name (str, optional): Name of specific sheet to process. If None, processes all sheets
    output_dir (str, optional): Directory where output file will be saved. Default is 'output'

    The function looks for cells with color #B987A6 (questions) and #65B0A4 (answers)
    """
    try:
        # Ensure output directory exists
        ensure_output_directory(output_dir)

        # Load workbook and determine sheets to process
        workbook, sheets_to_process = load_workbook_and_get_sheets(
            excel_file_path, sheet_name
        )

        # Process each sheet
        for current_sheet_name in sheets_to_process:
            worksheet = workbook[current_sheet_name]
            logger.info(f"Processing sheet: {current_sheet_name}")

            # Process questions and answers
            answer_columns, question_columns, max_text_lengths, rows_to_adjust = (
                process_sheet_questions(worksheet, question_processor_function)
            )

            # Adjust column widths
            adjust_column_widths(worksheet, answer_columns, max_text_lengths)

            # Adjust row heights
            adjust_row_heights(worksheet, rows_to_adjust)

            logger.info(f"Completed processing sheet: {current_sheet_name}")

        # Create a unique filename using timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        original_filename = os.path.splitext(os.path.basename(excel_file_path))[0]
        new_filename = f"{original_filename}_answers_filled.xlsx"

        # Save the workbook
        output_file = os.path.join(output_dir, new_filename)
        workbook.save(output_file)
        logger.info(f"Processed all sheets in {excel_file_path}")
        logger.info(f"Saved results to {output_file}")
        logger.info("Adjusted column widths and row heights for better readability")

        return output_file
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        raise
